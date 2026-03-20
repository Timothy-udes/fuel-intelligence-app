import streamlit as st
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

st.title("Global Fueling Report Formatter (Smart Professional)")

uploaded_file = st.file_uploader("Upload Excel Workbook", type=["xlsx"])

def enhance_format(ws):

    max_row = ws.max_row
    max_col = ws.max_column

    # COLORS
    title_fill = PatternFill(start_color="1F4E78", fill_type="solid")
    header_fill = PatternFill(start_color="D9E1F2", fill_type="solid")
    total_fill = PatternFill(start_color="FCE4D6", fill_type="solid")

    # BORDERS
    thick = Side(style='medium')
    border = Border(left=thick, right=thick, top=thick, bottom=thick)

    # ===== DETECT HEADER ROW =====
    header_row = 2

    # ===== FIND GRAND TOTAL =====
    grand_total_row = None
    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            val = ws.cell(row=row, column=col).value
            if val and str(val).strip().upper() == "GRAND TOTAL":
                grand_total_row = row
                break
        if grand_total_row:
            break

    # ===== IDENTIFY LITRES & COST COLUMNS =====
    litres_col = None
    cost_col = None

    for col in range(1, max_col + 1):
        header = str(ws.cell(row=header_row, column=col).value).upper()

        if "LITRE" in header:
            litres_col = col
        if "COST" in header:
            cost_col = col

    # ===== CREATE GRAND TOTAL IF MISSING =====
    if not grand_total_row and litres_col and cost_col:

        total_litres = 0
        total_cost = 0

        for row in range(header_row + 1, max_row + 1):
            try:
                total_litres += float(ws.cell(row=row, column=litres_col).value or 0)
                total_cost += float(ws.cell(row=row, column=cost_col).value or 0)
            except:
                pass

        grand_total_row = max_row + 1

        # Insert GRAND TOTAL label
        ws.cell(row=grand_total_row, column=2).value = "GRAND TOTAL"

        # Insert totals
        ws.cell(row=grand_total_row, column=litres_col).value = total_litres
        ws.cell(row=grand_total_row, column=cost_col).value = total_cost

    # ===== TITLE =====
    if ws.cell(row=1, column=1).value:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
        c = ws.cell(row=1, column=1)
        c.fill = title_fill
        c.font = Font(bold=True, size=16, color="FFFFFF")
        c.alignment = Alignment(horizontal="center")

    # ===== HEADERS =====
    for col in range(1, max_col + 1):
        c = ws.cell(row=header_row, column=col)
        if c.value:
            c.fill = header_fill
            c.font = Font(bold=True)
            c.alignment = Alignment(horizontal="center")
            c.border = border

    # ===== NUMBER FORMAT =====
    for col in [litres_col, cost_col]:
        if col:
            for row in range(header_row + 1, ws.max_row + 1):
                ws.cell(row=row, column=col).number_format = "#,##0.00"

    # ===== STYLE GRAND TOTAL =====
    if grand_total_row:
        for col in range(1, max_col + 1):
            c = ws.cell(row=grand_total_row, column=col)
            c.fill = total_fill
            c.font = Font(bold=True)
            c.border = border

    # ===== BORDERS =====
    for row in ws.iter_rows(min_row=header_row, max_row=ws.max_row, min_col=1, max_col=max_col):
        for c in row:
            c.border = border

    # ===== AUTO WIDTH =====
    for col in range(1, max_col + 1):
        ws.column_dimensions[get_column_letter(col)].width = 20


if uploaded_file:

    if st.button("Enhance & Fix Grand Totals"):

        wb = load_workbook(uploaded_file)

        for sheet in wb.sheetnames:
            enhance_format(wb[sheet])

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        st.download_button(
            label="Download Final Professional Report",
            data=output,
            file_name="Final_Enhanced_Report.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )