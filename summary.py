import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter

st.title("Global Fueling Report Builder (Professional)")

uploaded_files = st.file_uploader(
    "Upload Excel Workbooks",
    type=["xlsx"],
    accept_multiple_files=True
)

def format_sheet(ws, df, title, month, year, border, include_percentage=False):
    max_col = df.shape[1]
    max_row = len(df) + 2

    # TITLE IN UPPERCASE
    full_title = f"{title} - {month or ''} {year or ''}".upper()
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
    header_cell = ws.cell(row=1, column=1)
    header_cell.value = full_title
    header_cell.alignment = Alignment(horizontal="center")
    header_cell.font = Font(size=14, bold=True)

    # HEADERS
    for col in range(1, max_col + 1):
        cell = ws.cell(row=2, column=col)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    ws.freeze_panes = "A3"

    # NUMBER FORMAT
    for col_name in ["LITRES", "FUEL COST"]:
        if col_name in df.columns:
            col_index = df.columns.get_loc(col_name) + 1
            for row in range(3, max_row + 1):
                ws.cell(row=row, column=col_index).number_format = "#,##0.00"

    if include_percentage and "PERCENTAGE" in df.columns:
        col_index = df.columns.get_loc("PERCENTAGE") + 1
        for row in range(3, max_row + 1):
            ws.cell(row=row, column=col_index).number_format = "0.00%"

    # BOLD GRAND TOTAL
    for col in range(1, max_col + 1):
        if ws.cell(row=max_row, column=2).value == "GRAND TOTAL":
            ws.cell(row=max_row, column=col).font = Font(bold=True, size=12)

    # BORDERS
    for row in ws.iter_rows(min_row=2, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row:
            cell.border = border

    # AUTO WIDTH
    for i in range(1, max_col + 1):
        col_letter = get_column_letter(i)
        max_length = 0
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=i, max_col=i):
            for cell in row:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length + 3


if uploaded_files:

    sheet_data = {}
    detected_month = None
    detected_year = None
    sheet_name_set = set()

    for file in uploaded_files:

        df = pd.read_excel(file)
        df.columns = df.columns.astype(str).str.strip().str.upper()

        # STANDARDIZE COLUMNS
        rename_map = {
            "VEHICLE TYPE": "VEHICLE MAKE",
            "BRAND": "VEHICLE MAKE",
            "COST OF FUEL": "FUEL COST",
            "DRIVER": "DRIVERS NAME",
            "DRIVER'S NAME": "DRIVERS NAME",
            "FILLING STATION": "FUELING STATION",
            "FUELLING STATION": "FUELING STATION",
            "LITERS": "LITRES",
            "MILAGE": "MILEAGE"
        }
        df.rename(columns={k: v for k, v in rename_map.items() if k in df.columns}, inplace=True)

        # STANDARDIZE CELL VALUES
        replace_map = {
            "Exxon Mobil": "EXXON MOBIL",
            "HSE": "HSE UNIT",
            "Ikeja Electric PLC": "Ikeja Electric Plc",
            "MASTER CARD": "MASTERCARD",
            "SUPPLY CHAIN MGT.": "Supply Chain Management",
            "Supply Chain Managment": "Supply Chain Management"
        }

        for col in df.select_dtypes(include=['object']).columns:
            df[col] = df[col].replace(replace_map)

        # UNIQUE SHEET NAME
        name = file.name.replace(".xlsx", "")[:31]
        while name in sheet_name_set:
            name = name[:28] + "_1"
        sheet_name_set.add(name)

        # CLEAN ROWS
        df = df.dropna(how='all')
        if len(df) > 0:
            df = df.iloc[:-1].reset_index(drop=True)

        # DETECT MONTH/YEAR
        for col in df.columns:
            try:
                temp_col = pd.to_datetime(df[col], errors="coerce")
                if temp_col.notna().any():
                    detected_month = temp_col.dt.strftime("%B").dropna().iloc[0]
                    detected_year = temp_col.dt.strftime("%Y").dropna().iloc[0]
                    break
            except:
                continue

        sheet_data[name] = df

    st.success(f"{len(sheet_data)} workbooks loaded")

    if st.button("Generate Global Fueling Report"):

        output = BytesIO()
        all_merged = pd.DataFrame()
        thin = Side(style='thin')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        with pd.ExcelWriter(output, engine="openpyxl") as writer:

            # ===================== INDIVIDUAL SHEETS =====================
            for sheet_name, df in sheet_data.items():
                for col in ["LITRES", "FUEL COST"]:
                    if col in df.columns:
                        df[col] = pd.to_numeric(df[col], errors="coerce")
                df.to_excel(writer, sheet_name=sheet_name, startrow=1, index=False)
                ws = writer.sheets[sheet_name]
                format_sheet(ws, df, sheet_name, detected_month, detected_year, border)
                all_merged = pd.concat([all_merged, df], ignore_index=True)

            # ===================== ALL MERGED =====================
            all_merged.to_excel(writer, sheet_name="ALL MERGED REPORT", startrow=1, index=False)
            ws = writer.sheets["ALL MERGED REPORT"]
            format_sheet(ws, all_merged, "ALL MERGED REPORT", detected_month, detected_year, border)

            # ===================== TOTAL BY CONTRACT =====================
            total_by_contract = all_merged.groupby("CONTRACT", as_index=False).agg({
                "LITRES": "sum",
                "FUEL COST": "sum"
            }).sort_values(by="CONTRACT")
            total_by_contract.insert(0, "S/N", range(1, len(total_by_contract)+1))
            grand_total = pd.DataFrame({
                "S/N": [""],
                "CONTRACT": ["GRAND TOTAL"],
                "LITRES": [total_by_contract["LITRES"].sum()],
                "FUEL COST": [total_by_contract["FUEL COST"].sum()]
            })
            total_by_contract = pd.concat([total_by_contract, grand_total], ignore_index=True)
            total_by_contract.to_excel(writer, sheet_name="TOTAL BY CONTRACT", startrow=1, index=False)
            ws = writer.sheets["TOTAL BY CONTRACT"]
            format_sheet(ws, total_by_contract, "TOTAL BY CONTRACT", detected_month, detected_year, border)

            # ===================== PERCENTAGE =====================
            pct_df = total_by_contract[total_by_contract["CONTRACT"] != "GRAND TOTAL"].copy()
            total_cost = pct_df["FUEL COST"].sum()
            pct_df["PERCENTAGE"] = pct_df["FUEL COST"] / total_cost
            grand_pct = pd.DataFrame({
                "S/N": [""],
                "CONTRACT": ["GRAND TOTAL"],
                "LITRES": [pct_df["LITRES"].sum()],
                "FUEL COST": [total_cost],
                "PERCENTAGE": [1]
            })
            pct_df = pd.concat([pct_df, grand_pct], ignore_index=True)
            pct_df.to_excel(writer, sheet_name="PERCENTAGE", startrow=1, index=False)
            ws = writer.sheets["PERCENTAGE"]
            format_sheet(ws, pct_df, "PERCENTAGE ANALYSIS", detected_month, detected_year, border, include_percentage=True)

            # ===================== SUMMARY =====================
            summary_df = all_merged[all_merged["CONTRACT"] != "GRAND TOTAL"].copy()
            contracts = summary_df["CONTRACT"].unique()
            sources = summary_df["SOURCE"].unique()

            # Create summary table
            summary_table = pd.DataFrame({"S/N": range(1, len(contracts)+1), "CONTRACT": contracts})

            # Add SUM OF LITRES and SUM OF FUEL COST per SOURCE
            for src in sources:
                src_data = summary_df[summary_df["SOURCE"]==src]
                litres_sum = []
                fuel_sum = []
                for c in contracts:
                    c_data = src_data[src_data["CONTRACT"]==c]
                    litres_sum.append(c_data["LITRES"].sum() if not c_data.empty else 0)
                    fuel_sum.append(c_data["FUEL COST"].sum() if not c_data.empty else 0)
                summary_table[(src,"SUM OF LITRES")] = litres_sum
                summary_table[(src,"SUM OF FUEL COST")] = fuel_sum

            # TOTAL column
            summary_table[("TOTAL","SUM OF LITRES")] = summary_table[[ (src,"SUM OF LITRES") for src in sources]].sum(axis=1)
            summary_table[("TOTAL","SUM OF FUEL COST")] = summary_table[[ (src,"SUM OF FUEL COST") for src in sources]].sum(axis=1)

            # GRAND TOTAL row
            grand_total_row = pd.DataFrame({
                "S/N":[""],
                "CONTRACT":["GRAND TOTAL"],
                **{(src,"SUM OF LITRES"): [summary_table[(src,"SUM OF LITRES")].sum()] for src in sources},
                **{(src,"SUM OF FUEL COST"): [summary_table[(src,"SUM OF FUEL COST")].sum()] for src in sources},
                ("TOTAL","SUM OF LITRES"):[summary_table[("TOTAL","SUM OF LITRES")].sum()],
                ("TOTAL","SUM OF FUEL COST"):[summary_table[("TOTAL","SUM OF FUEL COST")].sum()]
            })
            summary_table = pd.concat([summary_table, grand_total_row], ignore_index=True)

            # Write to Excel
            summary_table.to_excel(writer, sheet_name="SUMMARY", startrow=1, index=False)
            ws = writer.sheets["SUMMARY"]

            # MERGE SOURCE HEADERS
            col_index = 3
            for src in sources:
                ws.merge_cells(start_row=1, start_column=col_index, end_row=1, end_column=col_index+1)
                ws.cell(row=1, column=col_index).value = src.upper()
                ws.cell(row=1, column=col_index).alignment = Alignment(horizontal="center")
                col_index += 2
            # Merge TOTAL
            ws.merge_cells(start_row=1, start_column=col_index, end_row=1, end_column=col_index+1)
            ws.cell(row=1, column=col_index).value = "TOTAL"
            ws.cell(row=1, column=col_index).alignment = Alignment(horizontal="center")

            # SECOND ROW HEADERS (SUM OF LITRES / SUM OF FUEL COST)
            for col in range(3, ws.max_column+1):
                cell = ws.cell(row=2, column=col)
                cell.value = "SUM OF LITRES" if col%2!=0 else "SUM OF FUEL COST"
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")

            # Format S/N and CONTRACT columns
            ws.cell(row=2, column=1).value = "S/N"
            ws.cell(row=2, column=2).value = "CONTRACT"
            ws.cell(row=2, column=1).font = Font(bold=True)
            ws.cell(row=2, column=2).font = Font(bold=True)
            ws.cell(row=2, column=1).alignment = Alignment(horizontal="center")
            ws.cell(row=2, column=2).alignment = Alignment(horizontal="center")

            # Borders & Auto width
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for cell in row:
                    cell.border = border

            for i in range(1, ws.max_column+1):
                col_letter = get_column_letter(i)
                max_length = 0
                for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=i, max_col=i):
                    for cell in row:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                ws.column_dimensions[col_letter].width = max_length + 3

        output.seek(0)
        st.download_button(
            label="Download Professional Report",
            data=output,
            file_name="Global_Fueling_Report_Final.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )